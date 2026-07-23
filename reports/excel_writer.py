import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
def clean(name):
    for b in ["\\","/","*","[","]",":","?"]: name=name.replace(b,"_")
    return name[:31]
def write_excel(output_file, data):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet, rows in data.items():
            df=pd.DataFrame(rows if rows else [{"Info":"No data found or no permission"}])
            df.to_excel(writer, sheet_name=clean(sheet), index=False)
        wb=writer.book
        for ws in wb.worksheets:
            ws.freeze_panes="A2"
            for c in ws[1]:
                c.font=Font(bold=True,color="FFFFFF")
                c.fill=PatternFill("solid", fgColor="1F4E78")
                c.alignment=Alignment(horizontal="center")
            for col in ws.columns:
                width=12
                letter=get_column_letter(col[0].column)
                for cell in col:
                    try: width=max(width, min(len(str(cell.value)),60))
                    except Exception: pass
                ws.column_dimensions[letter].width=width+2
